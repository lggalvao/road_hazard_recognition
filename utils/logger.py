import matplotlib.pyplot as plt
from pathlib import Path
from dataclasses import asdict
import wandb
import os

from pathlib import Path
import matplotlib.pyplot as plt
from sklearn.metrics import ConfusionMatrixDisplay
from utils.timing import timeit
import logging
import pandas as pd
import json
from typing import Tuple
import warnings
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


unwanted_keys = [
    'system_device',
    'system_seed',
    'data_input_img_resize',
    'data_train_videos_number',
    'data_test_videos_number',
    'data_train_video_samples_per_class',
    'data_train_total_video_samples',
    'data_test_samples_per_calss',
    'data_train_sequences_per_class', 
    'data_test_total_video_samples',
    'data_test_sequences_per_class',
    'data_img_shape',
    'data_load_normalization',
    'data_split_dataset',
    'data_dataset_folder_path',
    'data_train_csv_set_output_path',
    'data_test_csv_set_output_path',
    'data_dataset_event_time_csv_file_path',
    'data_dataset_csv_file_path',
    'data_num_workers',
    'data_no_hazard_samples_train_csv_file_path',
    'data_no_hazard_samples_test_csv_file_path',
    'model_classes_name',
    'logging_log_dir',
    'logging_results_csv_file_path',
    'logging_wandb_entity',
    'logging_wandb_project',
    'logging_wandb_run_name',
    'data_test_video_samples_per_class',
    'system_root',
    'data_saved_dataloader',
    'data_num_of_input_img',
    'logging_file_name',
    'logging_file_name',
    'logging_pred_save_frame_flag',
    'debugging_transformed_input_img_dir',
    '',
    
]


# Optional: colored output
class ColoredFormatter(logging.Formatter):
    """Adds color to log levels for console output."""

    COLORS = {
        'DEBUG': '\033[37m',    # White
        'INFO': '\033[36m',     # Cyan
        'WARNING': '\033[33m',  # Yellow
        'ERROR': '\033[31m',    # Red
        'CRITICAL': '\033[41m', # Red background
    }
    RESET = '\033[0m'

    def format(self, record):
        msg = super().format(record)
        color = self.COLORS.get(record.levelname, self.RESET)
        return f"{color}{msg}{self.RESET}"


def setup_logging(log_file: str = None, level=logging.INFO):
    """Sets up logging to console (with colors) and optional file. Warnings are ignored."""
    
    # Ignore warnings globally
    warnings.filterwarnings("ignore")
    
    logger = logging.getLogger()
    logger.setLevel(level)

    # Colored formatter for console
    console_formatter = ColoredFormatter('[%(asctime)s] %(levelname)s: %(message)s',
                                         '%Y-%m-%d %H:%M:%S')

    # Console handler
    ch = logging.StreamHandler()
    ch.setLevel(level)
    ch.setFormatter(console_formatter)
    logger.addHandler(ch)

    # File handler (plain, no color)
    if log_file:
        file_formatter = logging.Formatter('[%(asctime)s] %(levelname)s: %(message)s',
                                           '%Y-%m-%d %H:%M:%S')
        fh = logging.FileHandler(log_file, mode='w')
        fh.setLevel(level)
        fh.setFormatter(file_formatter)
        logger.addHandler(fh)

    return logger


@timeit
def log_results(
    epoch: int,
    metrics: dict,
    run_wandb=None,
    cm=None,
    classification_report: str = None,
    phase: str = "train",
    log_file_path: str = None,
    classes: list = None):
    """
    Log metrics to wandb, optionally save to file, and optionally plot confusion matrix.

    Args:
        epoch (int): Current epoch number.
        metrics (dict): Dictionary containing metrics such as loss, accuracy, precision, recall, f1.
        run_wandb: Wandb run object. If None, skips wandb logging.
        cm (optional): Confusion matrix to plot/save.
        classification_report (str, optional): Sklearn classification report.
        phase (str): "train", "val", or "test".
        log_file_path (str, optional): Path to save metrics and figures.
        classes (list, optional): List of class names for confusion matrix.
    """

    # -------------------------
    # 1. WandB Logging
    # -------------------------
    if run_wandb is not None and phase != "test" :
        # Use .get to avoid KeyError if a metric is missing
        wandb_metrics = {f"{phase}_{k}": metrics.get(k, 0) for k in metrics}
        run_wandb.log(wandb_metrics, step=epoch)

    # -------------------------
    # 2. File Logging
    # -------------------------
    if log_file_path:
        log_path = Path(log_file_path)
        log_path.mkdir(parents=True, exist_ok=True)
        log_file = log_path / "log.txt"

        # Different formatting for test vs train/val
        if phase == "test" and classification_report:
            with log_file.open("a") as f:
                f.write(f"\nTest Accuracy: {metrics.get('acc', 0):.3f}\n")
                f.write("Classification Report:\n")
                f.write(classification_report + "\n")
        else:
            # For train/val, simple epoch logging
            with log_file.open("a") as f:
                f.write(f"{epoch:.3f}\t{phase}\tacc:{metrics.get('acc', 0):.3f}\tloss:{metrics.get('loss', 0):.3f}\tf1_macro:{metrics.get('f1_macro', 0):.3f}\n")

    # -------------------------
    # 3. Confusion Matrix Figure (Test Only)
    # -------------------------
    if phase == "test" and cm is not None and classes is not None:
        cm_display = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=classes)
        
        df_cm = pd.DataFrame(cm, index=classes, columns=classes)
        df_cm.to_csv(log_path / "confusion_matrix.csv")
        
        #fig, ax = plt.subplots(figsize=(9, 7), dpi=120)
        fig, ax = plt.subplots(figsize=(10, 9), dpi=120)
        
        cm_display.plot(
            ax=ax,
            cmap='Blues',
            values_format=".2f",
            colorbar=True
        )

        ax.set_title("Confusion Matrix", fontsize=14, pad=12)
        ax.set_xlabel("Predicted Label", fontsize=11)
        ax.set_ylabel("True Label", fontsize=11)
        
        # Improve figure layout
        plt.xticks(rotation=90, ha='center')
        plt.tick_params(axis='x', labelsize=10)
        plt.tick_params(axis='y', labelsize=10)
        plt.tight_layout()

        if log_file_path:
            plt.savefig(log_path / "test_ConfusionMatrix.pdf", dpi=300, format='pdf')

        plt.close('all')


def initialize_wandb(cfg, log_file_path):
    return wandb.init(
        # Set the wandb entity where your project will be logged (generally your team name)
        entity=cfg.logging.wandb_entity,
        # Set the wandb project where this run will be logged
        project=cfg.logging.wandb_project,
        name=log_file_path,
        # Track hyperparameters and run metadata
        config={
                "model": cfg.model.model
               }
    )

@timeit
def draw_curve(current_epoch, x_epoch, epoch_y_loss, epoch_y_acc, log_file_path, file_number):
    """
    Draws and saves loss and accuracy curves.

    Args:
        current_epoch (int): Current epoch number.
        x_epoch (list): List of epoch indices.
        epoch_y_loss (dict): {"train": [...], "val": [...]} losses.
        epoch_y_acc (dict): {"train": [...], "val": [...]} accuracies.
        log_file_path (str or Path): Directory where plots are saved.
        file_number (int): Run or file identifier.
    """
    # Ensure path is Path object
    log_file_path = Path(log_file_path)

    # Append current epoch
    x_epoch.append(current_epoch)

    # Create figure and axes
    fig, (ax0, ax1, ax2) = plt.subplots(1, 3, figsize=(13, 12))

    # Loss curves
    ax0.plot(x_epoch, epoch_y_loss['train'], 'bo-', label='train')
    ax0.plot(x_epoch, epoch_y_loss['val'], 'ro-', label='val')
    ax0.set_title("Loss")
    ax0.legend()

    # Accuracy curves
    ax1.plot(x_epoch, epoch_y_acc['train'], 'bo-', label='train')
    ax1.plot(x_epoch, epoch_y_acc['val'], 'ro-', label='val')
    ax1.set_title("Accuracy")
    ax1.legend()

    # Third subplot reserved (empty for now)
    ax2.set_title("")

    # Save plots
    output_path_1 = log_file_path / f"lossGraphs_{file_number}.jpg"
    output_path_2 = log_file_path.parent / f"lossGraphs_{file_number}.jpg"

    fig.savefig(output_path_1)
    fig.savefig(output_path_2)

    plt.close(fig)


def log_config_to_wandb(cfg):
    cfg_dict = asdict(cfg)
    for section, params in cfg_dict.items():
        if isinstance(params, dict):
            for key, value in params.items():
                wandb.config[f"{section}.{key}"] = value
        else:
            wandb.config[section] = params


def flatten_cfg(cfg) -> dict:
    """
    Recursively flatten a nested cfg object into a dictionary suitable for CSV.
    Nested attributes become keys like 'data_num_dynamic_features'.
    """
    flat_dict = {}

    def _flatten(obj, prefix=''):
        for k, v in vars(obj).items():
            if hasattr(v, '__dict__'):
                _flatten(v, prefix + k + '_')
            else:
                flat_dict[prefix + k] = v

    _flatten(cfg)
    return flat_dict


def record_parameters_and_results(cfg) -> Tuple[str, int]:
    """
    Create a unique output folder for the run, save config, and record parameters in CSV.

    Returns:
        log_file_path: Path to folder for this run
        file_number: Unique run number
    """

    # --- Create unique output folder ---
    file_number = 0
    while True:
        test_name = f"{cfg.logging.file_name}_{file_number}"
        log_file_path = os.path.join('./output/trained_models', test_name)
        try:
            os.makedirs(log_file_path)
        except FileExistsError:
            file_number += 1
        else:
            break

    print(f"Directory created for run: {log_file_path}")

    cfg.logging.test_name = test_name

    return log_file_path, file_number


def update_results_csv(cfg, metrics = None) -> pd.DataFrame:
    """
    Add or update a column in results_csv for the current run.
    Keeps all existing columns and appends new parameters if needed.
    """
    # Flatten config and remove unwanted keys
    flat_cfg = flatten_cfg(cfg)
    flat_cfg = {k: v for k, v in flat_cfg.items() if k not in unwanted_keys}

    if metrics:
        # Add metrics placeholders
        metrics_placeholders = {
            'f1_macro': round(metrics["f1_macro"], 3),
            'acc': round(metrics["acc"], 3),
            'precision_weighted': round(metrics["precision_weighted"], 3),
            'recall_weighted': round(metrics["recall_weighted"], 3),
            'f1_weighted': round(metrics["f1_weighted"], 3),
            'precision_macro': round(metrics["precision_macro"], 3),
            'recall_macro': round(metrics["recall_macro"], 3),
        }
        flat_cfg.update(metrics_placeholders)

    # --- Load existing CSV ---
    if os.path.exists(cfg.logging.results_csv_file_path):
        results_csv = pd.read_csv(cfg.logging.results_csv_file_path)
    else:
        results_csv = pd.DataFrame(columns=['Parameters'])

    # --- Ensure 'Parameters' column exists ---
    if 'Parameters' not in results_csv.columns:
        results_csv['Parameters'] = []

    # --- Add new run column if missing ---
    if cfg.logging.test_name not in results_csv.columns:
        results_csv[cfg.logging.test_name] = ""

    # --- Update or append parameter rows ---
    for key, value in flat_cfg.items():
        if key not in results_csv['Parameters'].values:
            # New parameter: append row
            new_row = pd.DataFrame({'Parameters': [key], cfg.logging.test_name: [str(value)]})
            results_csv = pd.concat([results_csv, new_row], ignore_index=True)
        else:
            # Existing parameter: update value
            results_csv.loc[results_csv['Parameters'] == key, cfg.logging.test_name] = str(value)

    # Identify columns
    param_col = results_csv.columns[0]               # First column is 'Parameters'
    
    xlsx_file_path = cfg.logging.results_csv_file_path.replace('.csv', '.xlsx')
    
    wb = load_workbook(xlsx_file_path)
    ws = wb.active
    
    # Extract last column dynamically
    current_run = results_csv.columns[-1]
    previous_run = results_csv.columns[-2]
    
    next_col = ws.max_column + 1
    
    # Write header
    ws.cell(row=1, column=next_col, value=current_run)
    
    # Define highlight style
    yellow_fill = PatternFill(start_color="FFFF00",
                            end_color="FFFF00",
                            fill_type="solid")
    
        # Ensure there are at least two run columns
    #if results_csv.shape[1] < 3:  # 'Parameters' + at least two runs
    #    print("Not enough columns to compare.")
    #    
    #    wb.save(xlsx_file_path)
    #    
    #    return results_csv
    
    # Write values + apply highlight condition
    for row_idx, (_, row) in enumerate(results_csv.iterrows(), start=2):
        
        value = row[current_run]
        ws.cell(row=row_idx, column=next_col, value=value)
    
        # Apply highlight condition
        if str(row[current_run]) != str(row[previous_run]):
            ws.cell(row=row_idx, column=next_col).fill = yellow_fill
    
    wb.save(xlsx_file_path)
    
    results_csv.to_csv(cfg.logging.results_csv_file_path, index=False)
    
    print(f"Column '{current_run}' added with conditional highlights.")

    return results_csv
